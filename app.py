"""
Flash Group - Airtime Order Automation
Streamlit app: upload client PO PDFs/Excel -> check stock on hand -> populate network order templates

Run:  streamlit run app.py
"""

import re
import io
import datetime
from collections import defaultdict

import openpyxl
import pandas as pd
import pdfplumber
import streamlit as st
from rapidfuzz import fuzz

# ---------------------------------------------------------------------------
# CONSTANTS
# ---------------------------------------------------------------------------
NETWORKS = ["MTN", "Telkom", "Vodacom", "Cell C"]
CONFIDENCE_LOW   = 55
CONFIDENCE_WARN  = 70

# ---------------------------------------------------------------------------
# PDF PARSING
# ---------------------------------------------------------------------------
def _parse_po_line(line: str):
    disc = re.search(r'\b(\d+\.\d{4})\b', line)
    if not disc:
        return None
    before = line[:disc.start()].strip()
    price = re.search(r'(\d[\d,]*\.\d+)\s*$', before)
    if not price:
        return None
    qty_desc = before[:price.start()].strip()
    parts = qty_desc.split()
    if len(parts) < 2:
        return None
    qty_parts, i = [], 0
    while i < len(parts):
        if re.match(r'^\d+$', parts[i]):
            qty_parts.append(parts[i])
            i += 1
            if i < len(parts) and re.match(r'^\d{3}$', parts[i]):
                qty_parts.append(parts[i])
                i += 1
            desc_parts = parts[i:]
            break
        else:
            desc_parts = parts[i:]
            break
    if not qty_parts or not desc_parts:
        return None
    return int(''.join(qty_parts)), ' '.join(desc_parts)


def extract_po_items(pdf_bytes: bytes) -> list:
    items, seen = [], set()
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            for line in (page.extract_text() or '').split('\n'):
                result = _parse_po_line(line.strip())
                if result:
                    qty, desc = result
                    key = desc.lower()
                    if key not in seen:
                        items.append({'qty': qty, 'description': desc})
                        seen.add(key)
    return items


def extract_po_items_excel(excel_bytes: bytes) -> list:
    """Read a client Excel PO (MTN or Telkom format).
    Header row has 'Product Code' in col A; description in col B; qty in col D.
    """
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    ws = wb.active
    items, seen = [], set()
    header_found = False
    desc_col = qty_col = None

    for row in ws.iter_rows(values_only=True):
        if row[0] == 'Product Code':
            header_found = True
            for i, v in enumerate(row):
                if v in ('Description', 'Stock Item'):
                    desc_col = i
                elif v == 'Qty':
                    qty_col = i
            continue
        if not header_found or desc_col is None or qty_col is None:
            continue
        desc = row[desc_col]
        qty  = row[qty_col]
        if not desc or qty is None:
            continue
        try:
            qty_int = int(qty)
        except (ValueError, TypeError):
            continue
        if qty_int <= 0:
            continue
        key = str(desc).strip().lower()
        if key not in seen:
            items.append({'qty': qty_int, 'description': str(desc).strip()})
            seen.add(key)
    return items


def guess_network(filename: str, items: list) -> str:
    fn = filename.upper()
    for net in NETWORKS:
        if net.upper() in fn or net.upper().replace(' ', '') in fn:
            return net
    descs = ' '.join(i['description'] for i in items).upper()
    if 'MOBILE E-VOUCHER' in descs or 'LTE ANYTIME' in descs:
        return 'Telkom'
    if 'PAY AS YOU GO' in descs or 'MTN' in descs:
        return 'MTN'
    return 'MTN'


# ---------------------------------------------------------------------------
# SOH LOADING
# ---------------------------------------------------------------------------
def load_soh(file_bytes: bytes) -> tuple:
    """Returns ({vendor: {product: qty}}, date_str).

    Filters to the MOST RECENT date in the report so that each new download
    reflects current stock, not a cumulative sum across all historical rows.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    def _to_date(v):
        if isinstance(v, (datetime.datetime, datetime.date)):
            return v.date() if isinstance(v, datetime.datetime) else v
        try:
            return datetime.date.fromisoformat(str(v)[:10])
        except Exception:
            return None

    # Pass 1: collect rows and find latest date
    data_rows = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == 'Date':
            header_found = True
            continue
        if not header_found:
            continue
        date_val = row[0]
        qty, product, vendor = row[3], row[8], row[11]  # col D=Qty, col I=Product, col L=Vendor
        if product and vendor and qty and isinstance(qty, (int, float)):
            data_rows.append((_to_date(date_val), int(qty), str(product), str(vendor)))

    if not data_rows:
        return {}, 'No data'

    valid_dates = [r[0] for r in data_rows if r[0] is not None]
    latest_date = max(valid_dates) if valid_dates else None
    date_str = latest_date.isoformat() if latest_date else 'Unknown'

    # Pass 2: sum only rows from the latest date
    soh = defaultdict(lambda: defaultdict(int))
    for (parsed_date, qty, product, vendor) in data_rows:
        if latest_date is None or parsed_date == latest_date:
            soh[vendor][product] += qty

    return {v: dict(p) for v, p in soh.items()}, date_str


# ---------------------------------------------------------------------------
# MATCHING ENGINE
# ---------------------------------------------------------------------------
_CONTENT_TYPES = ['tiktok', 'whatsapp', 'youtube', 'social', 'chat', 'freeme']

def _norm(text: str) -> str:
    return re.sub(r'\+', ' ', text)

def _sizes(text: str) -> set:
    raw = re.findall(r'\d+(?:\.\d+)?(?:mb|gb)', text.lower())
    result = set()
    for s in raw:
        s = re.sub(r'\.0+(gb|mb)$', r'\1', s)
        result.add(s)
    return result

def _rands(text: str) -> set:
    return set(re.findall(r'r(\d+)', text.lower()))

def _mins(text: str) -> set:
    return set(re.findall(r'(\d+)\s*min', text.lower()))

def _period(text: str):
    t = text.lower()
    if re.search(r'3.?day|3day', t):   return '3day'
    if re.search(r'14.?day|14day', t): return '14day'
    if re.search(r'20.?day|20day', t): return '20day'
    if re.search(r'31.?day|31day', t): return 'monthly'
    if re.search(r'hourly|1.?hour|1hr', t): return 'hourly'
    if re.search(r'once.?off|onceoff', t): return 'onceoff'
    if re.search(r'daily|24hr', t):    return 'daily'
    if re.search(r'weekend', t):       return 'weekend'
    if re.search(r'weekly|7.?day', t): return 'weekly'
    if re.search(r'monthly|30.?day|month', t): return 'monthly'
    return None


def _score(query: str, candidate: str) -> float:
    q, c = _norm(query), _norm(candidate)
    qs, qr, qm = _sizes(q), _rands(q), _mins(q)
    cs, cr, cm = _sizes(c), _rands(c), _mins(c)
    if qs and cs and not qs & cs:  return 0.0
    if qs and not cs:              return 0.0
    if qr and cr and not qr & cr: return 0.0
    if qm and cm and not qm & cm: return 0.0
    qp, cp = _period(q), _period(c)
    penalty = 0
    if qp and cp and qp != cp and cp != 'onceoff':
        penalty += 20
    if cp == 'onceoff' and qp != 'onceoff':
        penalty += 15
    if re.search(r'all.net.voice|voice\+', c.lower()) and not re.search(r'voice', q.lower()):
        penalty += 15
    for ct in _CONTENT_TYPES:
        if (ct in c.lower()) != (ct in q.lower()):
            penalty += 15
            break
    score = fuzz.token_set_ratio(q.lower(), c.lower()) - penalty
    if qs and qs == cs: score = min(100.0, score + 15)
    if qr and qr == cr: score = min(100.0, score + 15)
    if qm and qm == cm: score = min(100.0, score + 10)
    return float(score)


def find_best(query: str, candidates: list, threshold: float = 45.0):
    best, best_score = None, 0.0
    for c in candidates:
        s = _score(query, c)
        if s > best_score:
            best_score, best = s, c
    if best_score >= threshold:
        return best, best_score
    return None, 0.0


# ---------------------------------------------------------------------------
# TEMPLATE LOADERS
# ---------------------------------------------------------------------------
def load_mtn_template_rows(file_bytes: bytes) -> list:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), keep_vba=True)
    ws = wb['ORDER TEMPLATE - (FLASH MyTown)']
    rows = []
    for row_cells in ws.iter_rows():
        actual_row = row_cells[0].row
        values = tuple(c.value for c in row_cells)
        if values[0] and values[1] and values[0] != 'PART #':
            rows.append((actual_row, str(values[1])))
    return rows


def load_telkom_template_rows(file_bytes: bytes) -> list:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb['Vouchers']
    SKIP_STARTS = {
        'Sub-Total', 'Total', 'Signature', 'Fax', 'E-mail', 'Orders',
        'Ship to', 'Dealer Code', 'Purchase Order', 'Customer Name',
        'Contact', 'Expected', 'Ref', 'Device Price', 'Product  ',
        'Sim Cards', 'SmartBroadband', 'Internet Starter', 'New Free-Me',
        'Data Bundles ', 'LTE/LTE-A', 'WhatsApp Bundles', 'Once-Off',
        'TikTok Data', 'All-networks', 'Product  - Virtual',
    }
    SIM_KW = ['sim card', 'blister', 'starter pack', 'internet sim']
    rows = []
    for row_cells in ws.iter_rows():
        actual_row = row_cells[0].row
        values = tuple(c.value for c in row_cells)
        name = values[1]
        mat_no = values[2]
        if not name or not mat_no:
            continue
        sname = str(name).strip()
        if any(sname.startswith(s) or sname == s for s in SKIP_STARTS):
            continue
        if any(k in sname.lower() for k in SIM_KW):
            continue
        if sname.startswith(('Huawei', 'Samsung', 'Sub', 'Product')):
            continue
        rows.append((actual_row, sname))
    return rows


# ---------------------------------------------------------------------------
# CORE PROCESSING
# ---------------------------------------------------------------------------
def process_orders(po_files, soh_bytes, mtn_template_bytes, telkom_template_bytes) -> pd.DataFrame:
    soh, soh_date = load_soh(soh_bytes)
    mtn_tmpl_rows    = load_mtn_template_rows(mtn_template_bytes)    if mtn_template_bytes    else []
    telkom_tmpl_rows = load_telkom_template_rows(telkom_template_bytes) if telkom_template_bytes else []
    mtn_tmpl_names    = [d for _, d in mtn_tmpl_rows]
    telkom_tmpl_names = [d for _, d in telkom_tmpl_rows]

    records = []
    for pof in po_files:
        items   = extract_po_items_excel(pof['bytes']) if pof.get('is_excel') else extract_po_items(pof['bytes'])
        network = pof['network']
        soh_vendor = soh.get(network, {})
        soh_names  = list(soh_vendor.keys())

        if network == 'MTN':
            tmpl_names, tmpl_rows = mtn_tmpl_names, mtn_tmpl_rows
        elif network == 'Telkom':
            tmpl_names, tmpl_rows = telkom_tmpl_names, telkom_tmpl_rows
        else:
            tmpl_names, tmpl_rows = [], []

        tmpl_row_map = {d: r for r, d in tmpl_rows}

        for item in items:
            desc   = item['description']
            po_qty = item['qty']

            soh_match, soh_conf = find_best(desc, soh_names)
            soh_qty   = soh_vendor.get(soh_match, 0) if soh_match else 0
            shortfall = max(0, po_qty - soh_qty)
            order_qty = shortfall * 8  # replenishment multiplier: order 8× the net shortfall

            tmpl_match, tmpl_conf = find_best(soh_match or desc, tmpl_names)
            tmpl_row_num = tmpl_row_map.get(tmpl_match) if tmpl_match else None

            records.append({
                'PO File'         : pof['name'],
                'Network'         : network,
                'PO Description'  : desc,
                'PO Qty'          : po_qty,
                'SOH Product'     : soh_match or '? No match',
                'SOH Qty'         : soh_qty,
                'SOH Confidence'  : round(soh_conf),
                'Shortfall'       : shortfall,
                'Order Qty'       : order_qty,
                'Template Row'    : tmpl_row_num,
                'Template Product': tmpl_match or '? No match',
                'Tmpl Confidence' : round(tmpl_conf),
            })

    result = pd.DataFrame(records)
    result.attrs['soh_date'] = soh_date
    return result


# ---------------------------------------------------------------------------
# TEMPLATE POPULATION
# ---------------------------------------------------------------------------
def _build_desc_row_map(ws, desc_col: int) -> dict:
    result = {}
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            if cell.column == desc_col and cell.value:
                result[str(cell.value).strip()] = cell.row
    return result


def populate_mtn_template(df: pd.DataFrame, template_bytes: bytes) -> bytes:
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes), keep_vba=True)
    ws = wb['ORDER TEMPLATE - (FLASH MyTown)']
    for merge_range in list(ws.merged_cells.ranges):
        if merge_range.min_col <= 4 <= merge_range.max_col:
            ws.unmerge_cells(str(merge_range))
    desc_to_row = _build_desc_row_map(ws, desc_col=2)
    mtn_orders = df[(df['Network'] == 'MTN') & (df['Template Product'] != '? No match')]
    written, not_found = [], []
    for _, row in mtn_orders.iterrows():
        tmpl_desc = str(row['Template Product']).strip()
        excel_row = desc_to_row.get(tmpl_desc)
        qty = int(row['Order Qty'])
        if excel_row and qty > 0:
            ws.cell(row=excel_row, column=4).value = qty
            written.append((tmpl_desc, excel_row, qty))
        elif not excel_row:
            not_found.append((tmpl_desc, qty))
    st.session_state['_mtn_write_log'] = {'written': written, 'not_found': not_found}
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def populate_telkom_template(df: pd.DataFrame, template_bytes: bytes) -> bytes:
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    ws = wb['Vouchers']
    for merge_range in list(ws.merged_cells.ranges):
        if merge_range.min_col <= 5 <= merge_range.max_col:
            ws.unmerge_cells(str(merge_range))
    desc_to_row = _build_desc_row_map(ws, desc_col=3)
    tel_orders = df[(df['Network'] == 'Telkom') & (df['Template Product'] != '? No match')]
    for _, row in tel_orders.iterrows():
        tmpl_desc = str(row['Template Product']).strip()
        excel_row = desc_to_row.get(tmpl_desc)
        qty = int(row['Order Qty'])
        if excel_row and qty > 0:
            ws.cell(row=excel_row, column=5).value = qty
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# STREAMLIT UI
# ---------------------------------------------------------------------------
FLASH_GREEN = "#B2FA00"
FLASH_BLACK = "#0D0D0D"

_LOGO_URI = (
    "data:image/svg+xml;base64,PHN2ZyB3aWR0aD0iMjkwIiBoZWlnaHQ9IjI4NSIgdmlld0JveD0iMCAwIDI5MCAyODUi"
    "IGZpbGw9Im5vbmUiIHhtbG5zPSJodHRwOi8vd3d3LnczLm9yZy8yMDAwL3N2ZyI+PHBhdGggZD0iTTI4OS45NjUgMTQy"
    "LjVDMjg5Ljk2NSAyMjEuMTk5IDIyNS4wNTMgMjg1IDE0NC45ODMgMjg1QzY0LjkxMjQgMjg1IDAgMjIxLjE5OSAwIDE0"
    "Mi41QzAgNjMuODAwOSA2NC45MTI0IDAgMTQ0Ljk4MyAwQzIyNS4wNTMgMCAyODkuOTY1IDYzLjgwMDkgMjg5Ljk2NSAx"
    "NDIuNVoiIGZpbGw9IiNCMkZBMDAiLz48cGF0aCBkPSJNMjMyLjQwMyAxMTkuNTU1QzIyMy40NjQgMTEyLjczNiAyMTMu"
    "NjIgMTAyLjM4NCAyMTQuOTMyIDg4LjEzMDhDMjIxLjYyIDg2LjI1NyAyMjUuNDY0IDgyLjIwMjIgMjI3LjI3NyA3OC4z"
    "OTMyQzIyOS45OTYgNzIuODAyNiAyMjkuODcxIDY3LjA1ODMgMjI0LjcxNCA2MS4zNzU1QzIyNC4wNTggNjAuNjY5IDIy"
    "Mi44NyA2MS4xNjA1IDIyMi45MzMgNjIuMTEyN0MyMjMuMTUyIDY2LjI5MDQgMjIzLjM3MSA3My42NjI3IDIxMy44MDcg"
    "NzUuNTk3OUMyMDMuMjc1IDc3Ljc0ODEgMTg1LjE3OSA3NS4xNjc4IDE3OS4wODUgNjkuNjM4NkMxNzMuNDYgNjQuNTM5"
    "NSAxNzIuODY2IDU4LjQyNjYgMTc3Ljk2IDQ4LjkwNDFDMTc4LjQyOSA0OC4wNDQgMTc3LjQ5MSA0Ny4xMjI0IDE3Ni42"
    "MTYgNDcuNTIxOEMxNzEuOTI4IDQ5Ljc2NDIgMTYzLjM2NSA1NC4xODc1IDE2My41NTIgNjYuOTA0N0MxNjMuNjc3IDc1"
    "LjM1MjEgMTY4LjY3OCA4Mi4xNDA4IDE3NS41ODUgODQuOTA1NEMxNDIuNzY5IDk0Ljg1OCAxMjkuMjM3IDgyLjg3OCA5"
    "MS45ODMgODIuMzU1OEM2MC45ODAxIDgxLjkyNTcgNTMuNzI5NCAxMDcuMjk5IDU1LjQ0ODMgMTMzLjI4NkM1Ni4xMzU4"
    "IDE0My41NDYgNTcuMDEwOSAxNTUuNjc5IDU5LjEwNDkgMTYzLjk0MkM2MC43OTI1IDE3MC43IDY0LjYwNTQgMTc0LjYz"
    "MiA2NC4yNjE2IDE4My4yMDJDNjQuMDExNiAxOTAuMTE0IDYzLjA0MjggMTkyLjk0IDY0LjE2NzkgMTk5LjUxNEM2NC40"
    "MTc5IDIwMC44OTYgNjUuMjkzIDIwMS43ODcgNjYuNzYxOSAyMDIuMjc4QzcwLjU0MzUgMjAzLjU2OCA3NS4wMTI2IDIw"
    "Mi45MjMgNzYuODI1MyAyMDIuMzdDNzguNTEzIDIwMS44NzkgNzkuNzk0MyAyMDAuOTg4IDc5LjQ1MDYgMTk5LjUxNEM3"
    "Ni40MTkgMTg2LjU1MSA3My42MDYzIDE2OS4wMTEgNzYuNTc1MyAxNTcuMzY5Qzc2Ljk4MTYgMTU1LjcxIDc4LjQxOTIg"
    "MTU0LjA4MiA4MS4wMTMyIDE1NS40NjRDODUuMjAxMSAxNTcuNzM3IDg5LjU3NjUgMTYxLjIzOSA5NS4zMjcxIDE2Ny4x"
    "NjhDMTA0LjAxNSAxNzYuMTM3IDExMi4yMzUgMTkzLjc2OSAxMTMuNTE2IDIwNy4wN0MxMTMuNzA0IDIwOS4wOTggMTE0"
    "LjcwNCAyMTAuNzg3IDExNi40MjMgMjExLjM3MUMxMjEuMDQ4IDIxMi45OTkgMTI1Ljc2NyAyMTIuNDE1IDEyOC4yMzYg"
    "MjEwLjkxQzEyOS4yOTkgMjEwLjI2NSAxMjkuODYyIDIwOS4yODIgMTI5LjU4IDIwOC4xNzZDMTI2LjQyNCAxOTUuNTgy"
    "IDExNi4xNDIgMTgwLjQwNyAxMTUuMjM1IDE2Ny4wMTRDMTE1LjA0OCAxNjQuMjggMTE2LjAxNyAxNjQuMDM1IDExOC4y"
    "NjcgMTY0LjQwM0MxMjEuNzM2IDE2NC45ODcgMTI0LjQ1NSAxNjUuMjAyIDEyNi45ODYgMTY1LjMyNUMxMjcuODMgMTY1"
    "LjM1NSAxMjguNjc0IDE2NS43MjQgMTI5LjIwNSAxNjYuMzY5QzEzOS42NzUgMTc4Ljk2MyAxNTEuNDg5IDE5Ni44MSAx"
    "NTAuNjQ1IDIxNS4zNjRDMTUwLjYxNCAyMTYuNDA4IDE1MS4wODIgMjE3LjM2MSAxNTEuODk1IDIxOC4wMDZDMTU1LjQ4"
    "OSAyMjAuOTI0IDE1NS41NTIgMjI3LjQzNiAxNTYuMTQ1IDIzMy4zOTVDMTU2LjI3IDIzNC42NTUgMTU3LjA1MiAyMzUu"
    "NzYxIDE1OC4yMzkgMjM2LjE5MUMxNjIuNTgzIDIzNy44MTkgMTY4LjExNSAyMzguMDk1IDE3Mi43MDkgMjM2LjM0NEMx"
    "NzMuNTIyIDIzNi4wMzcgMTc0LjA1MyAyMzUuMjY5IDE3My45MjggMjM0LjQ0QzE3Mi4yMDkgMjIyLjczNiAxNjIuODAy"
    "IDIxMi45MzcgMTY1LjAyMSAxOTUuNDI4QzE2NS40OSAxOTEuNjE5IDE2OC41NTMgMTg2LjE1MSAxNzAuNjE1IDE3OS4w"
    "ODZDMTc1LjUyMiAxNjIuMzQ1IDE3NC4zMDMgMTU3LjAzMSAxNzkuMTE2IDE0Ny41MDhDMTgyLjIxIDE0MS4zNjUgMTg5"
    "LjI0MiAxMzMuNDQgMTk5LjAyNCAxMzguOTk5QzIwNi41MjUgMTQzLjIzOSAyMTEuNzQ0IDE0NC4xMjkgMjE3LjEyIDE0"
    "My40MjNDMjI2LjQ5NiAxNDIuMTk0IDIzMi41MjggMTM0LjIwNyAyMzQuNzE1IDEyNS44MjJDMjM1LjIxNSAxMjMuNDg3"
    "IDIzNC4zNCAxMjEuMDMgMjMyLjQwMyAxMTkuNTU1WiIgZmlsbD0iYmxhY2siLz48L3N2Zz4="
)
LOGO_LG = '<img src="' + _LOGO_URI + '" width="56" height="56" style="flex-shrink:0">'
LOGO_SM = '<img src="' + _LOGO_URI + '" width="40" height="40" style="flex-shrink:0">'

st.set_page_config(
    page_title="Flash Cellular - Order Automation",
    page_icon=_LOGO_URI,
    layout="wide",
)

st.markdown(f"""
<style>
  div.stButton > button[kind="primary"] {{
      background-color: {FLASH_GREEN} !important;
      color: {FLASH_BLACK} !important;
      border: none !important; font-weight: 700 !important;
  }}
  div.stButton > button[kind="primary"]:hover {{ background-color: #c8ff1a !important; }}
  div.stDownloadButton > button {{
      border: 2px solid {FLASH_GREEN} !important;
      color: {FLASH_GREEN} !important; font-weight: 600 !important;
  }}
  section[data-testid="stSidebar"] {{ border-right: 3px solid {FLASH_GREEN}; }}
  .fh {{ display:flex; align-items:center; gap:16px;
          padding-bottom:10px; border-bottom:3px solid {FLASH_GREEN}; margin-bottom:4px; }}
  .fh h1 {{ margin:0; font-size:2rem; font-weight:800; line-height:1.1; }}
  .badge {{ display:inline-block; background:{FLASH_GREEN}; color:{FLASH_BLACK};
      font-size:.75rem; font-weight:700; letter-spacing:.08em;
      text-transform:uppercase; padding:2px 10px; border-radius:20px; margin-top:4px; }}
</style>
""", unsafe_allow_html=True)

st.markdown(
    '<div class="fh">' + LOGO_LG + '<div><h1>Flash Order Automation</h1>'
    '<span class="badge">Cellular Team</span></div></div>',
    unsafe_allow_html=True,
)
st.caption("Upload client POs  ->  check stock on hand  ->  populate network order templates for review")

with st.sidebar:
    st.markdown(
        '<div style="display:flex;align-items:center;gap:10px;padding-bottom:12px;'
        f'border-bottom:2px solid {FLASH_GREEN};margin-bottom:8px;">'
        + LOGO_SM +
        '<div><div style="font-weight:800;font-size:1rem;">Flash Group</div>'
        f'<div style="font-size:.72rem;color:{FLASH_GREEN};font-weight:700;'
        'letter-spacing:.07em;text-transform:uppercase;">Cellular Team</div>'
        '</div></div>',
        unsafe_allow_html=True,
    )
    st.header("Reference Files")
    st.caption("Upload once per session - or re-upload when files are updated.")

    soh_file = st.file_uploader("Stock on Hand Report (.xlsx)", type=['xlsx'], key='soh')
    mtn_tmpl_file = st.file_uploader("MTN Order Template (.xlsm / .xlsx)", type=['xlsm', 'xlsx'], key='mtn_tmpl')
    telkom_tmpl_file = st.file_uploader("Telkom Order Template (.xlsx)", type=['xlsx'], key='telkom_tmpl')

    st.divider()
    st.header("Client Purchase Orders")
    po_uploads = st.file_uploader(
        "Client POs - PDF or Excel (one or more)",
        type=['pdf', 'xlsx', 'xls'],
        accept_multiple_files=True,
        key='pos',
    )

# Network selection for each PO
po_files_ready = []

if po_uploads:
    st.subheader("Step 1 - Confirm network for each PO")
    cols = st.columns(min(len(po_uploads), 3))
    for i, f in enumerate(po_uploads):
        raw = f.read()
        is_excel = f.name.lower().endswith(('.xlsx', '.xls'))
        preview_items = extract_po_items_excel(raw) if is_excel else extract_po_items(raw)
        auto_net = guess_network(f.name, preview_items)
        if st.session_state.get(f'prev_po_{i}') != f.name:
            st.session_state[f'net_{i}'] = auto_net
        st.session_state[f'prev_po_{i}'] = f.name
        with cols[i % len(cols)]:
            st.markdown(f"**{f.name}**")
            st.caption(f"{len(preview_items)} line items detected")
            chosen = st.selectbox("Network", NETWORKS, index=NETWORKS.index(auto_net), key=f'net_{i}')
        po_files_ready.append({'name': f.name, 'bytes': raw, 'network': chosen, 'is_excel': is_excel})

# Check required files
networks_in_pos   = {pof['network'] for pof in po_files_ready}
needs_mtn_tmpl    = 'MTN'    in networks_in_pos
needs_telkom_tmpl = 'Telkom' in networks_in_pos

missing = []
if not soh_file:                               missing.append("SOH Report")
if needs_mtn_tmpl    and not mtn_tmpl_file:    missing.append("MTN Template")
if needs_telkom_tmpl and not telkom_tmpl_file: missing.append("Telkom Template")
if not po_files_ready:                         missing.append("at least one client PO")

if missing:
    st.info(f"Upload to continue: {', '.join(missing)}")
    st.stop()

# Fingerprint uploaded files; clear stale results when anything changes
_fp_parts = [soh_file.name, str(soh_file.size)] if soh_file else ['', '0']
if mtn_tmpl_file:    _fp_parts += [mtn_tmpl_file.name,    str(mtn_tmpl_file.size)]
if telkom_tmpl_file: _fp_parts += [telkom_tmpl_file.name, str(telkom_tmpl_file.size)]
for pof in po_files_ready:
    _fp_parts += [pof['name'], str(len(pof['bytes']))]
_current_fp = '|'.join(_fp_parts)

if st.session_state.get('_upload_fp') != _current_fp:
    for _k in ('results_df', 'mtn_tmpl_bytes', 'telkom_tmpl_bytes', '_mtn_write_log', 'soh_date'):
        st.session_state.pop(_k, None)
    st.session_state['_upload_fp'] = _current_fp

if st.button("Analyse Orders", type="primary", use_container_width=True):
    soh_bytes_val         = soh_file.read()
    mtn_tmpl_bytes_val    = mtn_tmpl_file.read()    if mtn_tmpl_file    else None
    telkom_tmpl_bytes_val = telkom_tmpl_file.read() if telkom_tmpl_file else None
    with st.spinner("Parsing POs and matching products..."):
        df = process_orders(
            po_files=po_files_ready,
            soh_bytes=soh_bytes_val,
            mtn_template_bytes=mtn_tmpl_bytes_val,
            telkom_template_bytes=telkom_tmpl_bytes_val,
        )
    st.session_state['results_df']        = df
    st.session_state['soh_date']          = df.attrs.get('soh_date', 'Unknown')
    st.session_state['mtn_tmpl_bytes']    = mtn_tmpl_bytes_val
    st.session_state['telkom_tmpl_bytes'] = telkom_tmpl_bytes_val

if 'results_df' not in st.session_state:
    st.stop()

df = st.session_state['results_df'].copy()

tab1, tab2, tab3 = st.tabs(["Stock Analysis", "Review & Adjust", "Download Templates"])

# Tab 1: Stock Analysis
with tab1:
    soh_date = st.session_state.get('soh_date', 'Unknown')
    st.info(f"SOH data as at: **{soh_date}** (only rows from this date are used — download a fresh VVS-SOH report each time)")

    n_items     = len(df)
    n_shortfall = int((df['Shortfall'] > 0).sum())
    total_sf    = int(df['Shortfall'].sum())
    n_low_conf  = int(((df['SOH Confidence'] < CONFIDENCE_WARN) | (df['Tmpl Confidence'] < CONFIDENCE_WARN)).sum())

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total PO Line Items", n_items)
    c2.metric("Items Needing Order", n_shortfall, delta=None if n_shortfall == 0 else f"{n_shortfall} shortfalls")
    c3.metric("Total Units to Order", total_sf)
    c4.metric("Low-Confidence Matches", n_low_conf, help="Review these in the next tab")

    st.divider()
    st.subheader("Full Analysis")

    def _colour_row(row):
        very_low = row['SOH Confidence'] < CONFIDENCE_LOW or row['Tmpl Confidence'] < CONFIDENCE_LOW
        low      = row['SOH Confidence'] < CONFIDENCE_WARN or row['Tmpl Confidence'] < CONFIDENCE_WARN
        sf       = row['Shortfall'] > 0
        if very_low:     return ['background-color: #f8d7da'] * len(row)   # red  — very low confidence
        if sf and low:   return ['background-color: #fff3cd'] * len(row)   # yellow — shortfall + low conf
        elif sf:         return ['background-color: #d1e7dd'] * len(row)   # green — shortfall to order
        elif low:        return ['background-color: #fff3cd'] * len(row)   # yellow — low confidence only
        return [''] * len(row)

    display_cols = [
        'Network', 'PO Description', 'PO Qty',
        'SOH Product', 'SOH Qty', 'SOH Confidence',
        'Shortfall', 'Template Product', 'Tmpl Confidence',
    ]
    styled = df[display_cols].style.apply(_colour_row, axis=1)
    st.dataframe(styled, use_container_width=True, height=600)
    st.caption("[G] Green = shortfall to order  [Y] Yellow = match confidence < 70% — review recommended  [R] Red = very low confidence < 55% — manual fix required")

# Tab 2: Review & Adjust
with tab2:
    st.subheader("Review matches and adjust quantities")
    st.caption("Change Order Qty to 0 to exclude an item. Fix wrong matches by editing the Template Product column.")

    show_filter = st.radio(
        "Show",
        ["Items to order (shortfall > 0)", "Low-confidence matches", "All items"],
        horizontal=True,
    )
    if show_filter == "Items to order (shortfall > 0)":
        edit_df = df[df['Shortfall'] > 0].copy()
    elif show_filter == "Low-confidence matches":
        edit_df = df[(df['SOH Confidence'] < CONFIDENCE_WARN) | (df['Tmpl Confidence'] < CONFIDENCE_WARN)].copy()
    else:
        edit_df = df.copy()

    if edit_df.empty:
        st.success("No items match this filter. Stock is sufficient for all ordered items!")
    else:
        edited = st.data_editor(
            edit_df[[
                'Network', 'PO File', 'PO Description', 'PO Qty',
                'SOH Product', 'SOH Qty', 'SOH Confidence',
                'Shortfall', 'Order Qty',
                'Template Product', 'Template Row', 'Tmpl Confidence',
            ]],
            use_container_width=True,
            num_rows="fixed",
            column_config={
                'Order Qty':       st.column_config.NumberColumn("Order Qty", min_value=0, step=1),
                'SOH Confidence':  st.column_config.ProgressColumn("SOH Conf", min_value=0, max_value=100, format="%d%%"),
                'Tmpl Confidence': st.column_config.ProgressColumn("Tmpl Conf", min_value=0, max_value=100, format="%d%%"),
            },
            disabled=['Network', 'PO File', 'PO Description', 'PO Qty',
                      'SOH Product', 'SOH Qty', 'SOH Confidence', 'Shortfall',
                      'Template Row', 'Tmpl Confidence'],
            key='editor',
        )

        if st.button("Save adjustments", type="secondary"):
            for col in ['Order Qty', 'Template Product']:
                df.loc[edited.index, col] = edited[col]
            st.session_state['results_df'] = df
            st.success("Adjustments saved. Switch to the Download tab.")

# Tab 3: Download Templates
with tab3:
    st.subheader("Download populated order templates")

    df_final = st.session_state['results_df']
    networks_in_order = df_final['Network'].unique().tolist()

    mtn_bytes_raw    = st.session_state.get('mtn_tmpl_bytes')
    telkom_bytes_raw = st.session_state.get('telkom_tmpl_bytes')

    if mtn_bytes_raw is None and mtn_tmpl_file:
        mtn_tmpl_file.seek(0)
        mtn_bytes_raw = mtn_tmpl_file.read()
    if telkom_bytes_raw is None and telkom_tmpl_file:
        telkom_tmpl_file.seek(0)
        telkom_bytes_raw = telkom_tmpl_file.read()

    if 'MTN' in networks_in_order and mtn_bytes_raw:
        mtn_orders = df_final[(df_final['Network'] == 'MTN') & (df_final['Order Qty'] > 0)]
        n_mtn = len(mtn_orders)
        total_mtn = int(mtn_orders['Order Qty'].sum())
        st.markdown("### MTN Order Template")
        st.markdown(f"**{n_mtn}** product(s) to order, **{total_mtn:,}** total units")
        if n_mtn > 0:
            st.dataframe(
                mtn_orders[['PO Description', 'SOH Product', 'PO Qty', 'SOH Qty', 'Order Qty', 'Template Product', 'Template Row']],
                use_container_width=True,
            )
            populated_mtn = populate_mtn_template(df_final, mtn_bytes_raw)
            log = st.session_state.get('_mtn_write_log', {})
            with st.expander("Write log", expanded=True):
                written   = log.get('written', [])
                not_found = log.get('not_found', [])
                if written:
                    st.success(f"Written to template ({len(written)} rows):")
                    st.dataframe(pd.DataFrame(written, columns=['Template Product', 'Excel Row', 'Qty']), use_container_width=True)
                if not_found:
                    st.error(f"Not found in template - description mismatch ({len(not_found)} rows):")
                    st.dataframe(pd.DataFrame(not_found, columns=['Template Product', 'Qty']), use_container_width=True)
                if not written and not not_found:
                    st.info("No quantities to write - SOH covers all ordered products.")
            st.download_button(
                label="Download MTN Order Template (.xlsm)",
                data=populated_mtn,
                file_name="MTN_Order_Template_POPULATED.xlsm",
                mime="application/vnd.ms-excel.sheet.macroEnabled.12",
            )
        else:
            st.success("No MTN stock needs to be ordered - SOH covers all client requests.")

    if 'Telkom' in networks_in_order and telkom_bytes_raw:
        tel_orders = df_final[(df_final['Network'] == 'Telkom') & (df_final['Order Qty'] > 0)]
        n_tel = len(tel_orders)
        total_tel = int(tel_orders['Order Qty'].sum())
        st.markdown("### Telkom Order Template")
        st.markdown(f"**{n_tel}** product(s) to order, **{total_tel:,}** total units")
        if n_tel > 0:
            st.dataframe(
                tel_orders[['PO Description', 'SOH Product', 'PO Qty', 'SOH Qty', 'Order Qty', 'Template Product']],
                use_container_width=True,
            )
            populated_telkom = populate_telkom_template(df_final, telkom_bytes_raw)
            st.download_button(
                label="Download Telkom Order Template (.xlsx)",
                data=populated_telkom,
                file_name="Telkom_Order_Template_POPULATED.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.success("No Telkom stock needs to be ordered - SOH covers all client requests.")

    unsupported_nets = [n for n in df_final['Network'].unique() if n not in ('MTN', 'Telkom')]
    if unsupported_nets:
        st.warning(
            f"**{', '.join(unsupported_nets)}** order template(s) are not yet supported. "
            "Products for these networks appear in the Stock Analysis tab, "
            "but template population and downloads are not available yet."
        )

    st.divider()
    st.subheader("Email drafts")
    st.caption("Copy these into your email client before sending the template.")
    networks_in_order = sorted(df_final[df_final["Order Qty"] > 0]["Network"].unique())
    for net in networks_in_order:
        net_orders = df_final[(df_final["Network"] == net) & (df_final["Order Qty"] > 0)]
        if net_orders.empty:
            continue
        n     = len(net_orders)
